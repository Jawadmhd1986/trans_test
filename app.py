from flask import Flask, render_template, request, jsonify, send_file, redirect, session, after_this_request
import pandas as pd
import os
from datetime import datetime, timedelta
from filelock import FileLock
from openpyxl import Workbook, load_workbook
from sqlalchemy import create_engine, Column, Integer, String, DateTime, ForeignKey, Boolean, Text, UniqueConstraint, Index, func, or_
from sqlalchemy.ext.declarative import declarative_base
from sqlalchemy.orm import sessionmaker, relationship
from werkzeug.security import generate_password_hash, check_password_hash
import json
import hashlib
import re
import pytz

app = Flask(__name__)
app.secret_key = os.environ.get("APP_SECRET", "dsv-stock-count-secret-key-2025")

# --- Timezone Configuration ---
ABU_DHABI_TZ = pytz.timezone("Asia/Dubai")

def abu_dhabi_now():
    """Returns the current time in Abu Dhabi timezone."""
    return datetime.now(ABU_DHABI_TZ)

def _no_cache(resp):
    resp.headers["Cache-Control"] = "no-store, no-cache, must-revalidate, max-age=0"
    resp.headers["Pragma"] = "no-cache"
    resp.headers["Expires"] = "0"
    return resp
app.config.update(
    SESSION_COOKIE_SAMESITE="Lax",
    SESSION_COOKIE_SECURE=True,
    SESSION_COOKIE_HTTPONLY=True,
    PERMANENT_SESSION_LIFETIME=timedelta(hours=8),
)

SESSION_TL_KEY = "tl_session"

def set_tl_session(tl_name, tl_display_name):
    session[SESSION_TL_KEY] = {"name": tl_name, "display_name": tl_display_name, "ts": abu_dhabi_now().isoformat()}

def require_tl():
    return bool(session.get(SESSION_TL_KEY))

# Database setup
engine = create_engine('sqlite:///line_count.db', echo=False)
Base = declarative_base()
SessionLocal = sessionmaker(autocommit=False, autoflush=False, bind=engine)

# Constants
LOCATIONS = ["KIZAD", "JEBEL_ALI"]
WAREHOUSES = {
    "KIZAD": ["KIZAD-W1"],
    "JEBEL_ALI": ["JA-W1", "JA-W2", "JA-W3"]
}

EXPORTS_DIR = os.path.join(os.getcwd(), "exports")
MDF_PATH = os.path.join(EXPORTS_DIR, "MDF.xlsx")
LOCK_PATH = os.path.join(EXPORTS_DIR, "MDF.lock")
COLUMNS = ["Date","Time","Location","Warehouse","CounterName","SKU","SerialOrCode","QTY","Source"]

# Models
class Line(Base):
    __tablename__ = 'lines'

    id = Column(Integer, primary_key=True)
    location = Column(String(50), nullable=False)
    warehouse = Column(String(50), nullable=False)
    line_code = Column(String(20), nullable=False)
    target_qty = Column(Integer, nullable=False)
    created_by_tl_norm = Column(String(120), nullable=True)
    created_at = Column(DateTime, default=abu_dhabi_now)
    updated_at = Column(DateTime, default=abu_dhabi_now, onupdate=abu_dhabi_now)

    assignments = relationship("Assignment", back_populates="line")
    scan_jobs = relationship("ScanJob", back_populates="line")

    __table_args__ = (Index('idx_line_warehouse', 'line_code', 'warehouse'),)

class Assignment(Base):
    __tablename__ = 'assignments'

    id = Column(Integer, primary_key=True)
    line_id = Column(Integer, ForeignKey('lines.id'), nullable=False)
    counter_name_1 = Column(String(100), nullable=False)
    counter_name_2 = Column(String(100), nullable=False)
    tl_name = Column(String(100), nullable=False)
    tl_pin_hash = Column(String(256), nullable=False)
    active = Column(Boolean, default=True)
    created_at = Column(DateTime, default=abu_dhabi_now)

    line = relationship("Line", back_populates="assignments")

class ScanJob(Base):
    __tablename__ = 'scan_jobs'

    id = Column(Integer, primary_key=True)
    line_id = Column(Integer, ForeignKey('lines.id'), nullable=False)
    status = Column(String(20), default='open')  # open, submitted, variance_approved
    opened_at = Column(DateTime, default=abu_dhabi_now)
    closed_at = Column(DateTime)
    opened_by = Column(String(100))

    line = relationship("Line", back_populates="scan_jobs")
    scans = relationship("Scan", back_populates="job")
    reconciliations = relationship("Reconciliation", back_populates="job")

class Scan(Base):
    __tablename__ = 'scans'

    id = Column(Integer, primary_key=True)
    job_id = Column(Integer, ForeignKey('scan_jobs.id'), nullable=False)
    line_id = Column(Integer, ForeignKey('lines.id'), nullable=False)
    counter_name = Column(String(100), nullable=False)
    sku = Column(String(100))
    serial_code = Column(String(200), nullable=False)
    qty = Column(Integer, default=1)
    source = Column(String(20), nullable=False)  # scan, manual
    created_at = Column(DateTime, default=abu_dhabi_now)

    job = relationship("ScanJob", back_populates="scans")

    __table_args__ = (
        Index('idx_job_serial', 'job_id', 'serial_code'),
    )

class Reconciliation(Base):
    __tablename__ = 'reconciliations'

    id = Column(Integer, primary_key=True)
    job_id = Column(Integer, ForeignKey('scan_jobs.id'), nullable=False)
    requested_by = Column(String(100), nullable=False)
    reason = Column(Text)
    previous_target = Column(Integer)
    new_target = Column(Integer)
    tl_approved_by = Column(String(100))
    approved_at = Column(DateTime)
    note = Column(Text)
    result = Column(String(20))  # approved_variance, edited_target, rejected

    job = relationship("ScanJob", back_populates="reconciliations")

class TLUser(Base):
    __tablename__ = "tl_users"

    id = Column(Integer, primary_key=True)
    name_norm = Column(String(120), nullable=False)      # normalized key
    display_name = Column(String(120), nullable=False)   # what you typed
    pin_hash = Column(String(255), nullable=True)
    role = Column(String(20), default='tl')
    created_at = Column(DateTime, server_default=func.now()) # This should also be updated if it represents creation time

    __table_args__ = (UniqueConstraint('name_norm', name='uq_tl_users_name_norm'),)

class ReconciliationQueue(Base):
    __tablename__ = 'reconciliation_queue'

    id = Column(Integer, primary_key=True)
    job_id = Column(Integer, ForeignKey('scan_jobs.id'), nullable=False)
    line_id = Column(Integer, ForeignKey('lines.id'), nullable=False)
    requested_by = Column(String(100), nullable=False)
    reason = Column(Text)
    scanned_total = Column(Integer, nullable=False)
    target_qty = Column(Integer, nullable=False)
    status = Column(String(20), default='pending')  # pending, approved, rejected
    tl_response = Column(Text)
    acknowledged = Column(Boolean, default=False)  # Track if counter acknowledged
    created_at = Column(DateTime, default=abu_dhabi_now)
    resolved_at = Column(DateTime)

    job = relationship("ScanJob")
    line = relationship("Line")

class ReconciliationRequest(Base):
    __tablename__ = 'reconciliation_requests'

    id = Column(Integer, primary_key=True)
    line_id = Column(Integer, ForeignKey('lines.id'), nullable=False)
    job_id = Column(Integer, ForeignKey('scan_jobs.id'), nullable=False)
    tl_name_norm = Column(String(120), nullable=False)
    requested_by = Column(String(100), nullable=False)
    requested_qty = Column(Integer, nullable=True)
    status = Column(String(20), default='pending')
    resolved_by = Column(String(100))
    resolved_at = Column(DateTime)
    created_at = Column(DateTime, default=abu_dhabi_now)

    line = relationship("Line")
    job = relationship("ScanJob")

class AuditLog(Base):
    __tablename__ = 'audit_log'

    id = Column(Integer, primary_key=True)
    actor = Column(String(100), nullable=False)
    action = Column(String(100), nullable=False)
    entity = Column(String(50), nullable=False)
    entity_id = Column(Integer)
    payload_json = Column(Text)
    created_at = Column(DateTime, default=abu_dhabi_now)

def init_db():
    """Initialize database and create tables"""
    Base.metadata.create_all(bind=engine)

    # Handle database migration for missing columns and indexes
    try:
        from sqlalchemy import text
        with engine.connect() as conn:
            # Check if acknowledged column exists
            result = conn.execute(text("PRAGMA table_info(reconciliation_queue)"))
            columns = [row[1] for row in result.fetchall()]

            if 'acknowledged' not in columns:
                print("Adding missing 'acknowledged' column to reconciliation_queue table...")
                conn.execute(text("ALTER TABLE reconciliation_queue ADD COLUMN acknowledged BOOLEAN DEFAULT 0"))
                conn.commit()
                print("Database migration completed successfully")

            # Check if created_by_tl_norm column exists in lines table
            result = conn.execute(text("PRAGMA table_info(lines)"))
            line_columns = [row[1] for row in result.fetchall()]

            if 'created_by_tl_norm' not in line_columns:
                print("Adding missing 'created_by_tl_norm' column to lines table...")
                conn.execute(text("ALTER TABLE lines ADD COLUMN created_by_tl_norm VARCHAR(120)"))
                conn.commit()
                print("Added created_by_tl_norm column")

            # Check if role column exists in tl_users table
            result = conn.execute(text("PRAGMA table_info(tl_users)"))
            tl_columns = [row[1] for row in result.fetchall()]

            if 'role' not in tl_columns:
                print("Adding missing 'role' column to tl_users table...")
                conn.execute(text("ALTER TABLE tl_users ADD COLUMN role VARCHAR(20) DEFAULT 'tl'"))
                conn.commit()
                print("Added role column")

            # Remove old unique constraint if it exists and add new composite index
            try:
                # Check if old constraint exists
                result = conn.execute(text("SELECT sql FROM sqlite_master WHERE type='index' AND name='unique_job_serial'"))
                if result.fetchone():
                    print("Removing old unique constraint...")
                    conn.execute(text("DROP INDEX IF EXISTS unique_job_serial"))
                    conn.commit()

                # Add composite unique index for proper duplicate checking
                conn.execute(text(
                    "CREATE UNIQUE INDEX IF NOT EXISTS ux_scans_job_sku_serial "
                    "ON scans (job_id, sku, serial_code)"
                ))
                conn.commit()
                print("Added composite unique index for scan duplicates")
            except Exception as idx_e:
                print(f"Index creation warning: {idx_e}")

    except Exception as e:
        print(f"Database migration warning: {e}")
        # Continue anyway as this is not critical

def ensure_mdf():
    """Ensure MDF.xlsx exists with proper headers"""
    os.makedirs(EXPORTS_DIR, exist_ok=True)
    if not os.path.exists(MDF_PATH):
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        ws.append(COLUMNS)
        wb.save(MDF_PATH)
        print(f"Created {MDF_PATH} with headers")

@app.before_request
def _boot():
    if not hasattr(app, '_initialized'):
        with app.app_context():
            init_db()
            ensure_mdf()
        app._initialized = True

def hash_pin(pin):
    """Hash PIN for secure storage"""
    return hashlib.sha256(pin.encode()).hexdigest()

def verify_pin(pin, pin_hash):
    """Verify PIN against hash"""
    return hashlib.sha256(pin.encode()).hexdigest() == pin_hash

def _norm(s):
    """Normalize string for case-insensitive comparison"""
    return (s or "").strip().casefold()

def _require_tl():
    return bool(session.get("tl_session"))

def _session_user():
    s = session.get("tl_session") or {}
    return s.get("display_name",""), s.get("name_norm","")

def _is_manager():
    _, nn = _session_user()
    db = SessionLocal()
    try:
        u = db.query(TLUser).filter_by(name_norm=nn).first()
        return bool(u and u.role == "manager")
    finally:
        db.close()

def norm_sku(s):
    """Normalize SKU for consistent storage"""
    return re.sub(r"[^A-Za-z0-9\-_]", "", (s or "").strip()).upper()

def norm_code(s):
    """Normalize serial/code for consistent storage"""
    return re.sub(r"[^A-Za-z0-9\-_]", "", (s or "").strip()).upper()

# Routes
@app.route('/signin')
def signin():
    return render_template('signin.html')

@app.route('/')
def home():
    return render_template('home.html', locations=LOCATIONS, warehouses=WAREHOUSES)

@app.route('/count')
def count():
    location = request.args.get('location')
    warehouse = request.args.get('warehouse')
    line_code = request.args.get('line')
    counter = request.args.get('counter')

    if not all([location, warehouse, line_code, counter]):
        return redirect('/')

    return render_template('count.html',
                         location=location,
                         warehouse=warehouse,
                         line_code=line_code,
                         counter=counter)

@app.route('/reconcile')
def reconcile_page():
    return render_template('reconcile_center.html')

@app.route('/log')
def log():
    db = SessionLocal()
    try:
        # Get completed jobs from database
        jobs = db.query(ScanJob).filter(
            ScanJob.status.in_(['submitted', 'variance_approved'])
        ).order_by(ScanJob.closed_at.desc()).all()

        job_data = []

        # Add database jobs
        for job in jobs:
            total_scans = db.query(Scan).filter(Scan.job_id == job.id).count()
            total_qty = db.query(func.sum(Scan.qty)).filter(Scan.job_id == job.id).scalar() or 0

            job_data.append({
                'line': job.line,
                'total_scans': total_scans,
                'total_qty': int(total_qty),
                'closed_at': job.closed_at,
                'status': job.status,
                'source': 'database'
            })

        # Also read historical data from MDF Excel file
        try:
            if os.path.exists(MDF_PATH):
                df = pd.read_excel(MDF_PATH)
                if not df.empty and len(df) > 0:
                    # Group by date, location, warehouse, and first counter name to identify unique jobs
                    historical_jobs = df.groupby(['Date', 'Location', 'Warehouse', 'CounterName']).agg({
                        'QTY': 'sum',
                        'SerialOrCode': 'count',
                        'Time': 'max'  # Get latest time for that job
                    }).reset_index()

                    # Add historical jobs that aren't already in the database
                    for _, row in historical_jobs.iterrows():
                        # Create a mock line object for display
                        class MockLine:
                            def __init__(self, location, warehouse, counter):
                                self.location = location
                                self.warehouse = warehouse
                                self.line_code = f"Historical-{counter}"
                                self.target_qty = "N/A"

                        # Parse the date and time
                        date_str = str(row['Date'])
                        time_str = str(row['Time'])

                        try:
                            if 'T' in date_str:  # ISO format
                                closed_at = datetime.fromisoformat(date_str.replace('Z', '+00:00'))
                            else:
                                # Try to combine date and time
                                datetime_str = f"{date_str} {time_str}"
                                closed_at = datetime.strptime(datetime_str, '%Y-%m-%d %H:%M:%S')
                        except:
                            closed_at = abu_dhabi_now()  # Fallback

                        job_data.append({
                            'line': MockLine(row['Location'], row['Warehouse'], row['CounterName']),
                            'total_scans': int(row['SerialOrCode']),
                            'total_qty': int(row['QTY']),
                            'closed_at': closed_at,
                            'status': 'historical',
                            'source': 'excel'
                        })
        except Exception as e:
            print(f"Error reading historical data: {e}")

        # Sort all jobs by closed_at date (newest first), handling None values
        job_data.sort(key=lambda x: x['closed_at'] or datetime.min, reverse=True)

        return render_template('log.html', jobs=job_data)
    finally:
        db.close()

# API Routes
@app.route('/api/tl/login', methods=['POST'])
def api_tl_login():
    """Validate TL PIN and create session"""
    data = request.get_json(force=True)
    tl_name = (data.get("tl_name") or "").strip()
    tl_pin = (data.get("tl_pin") or "").strip()
    tl_display_name = (data.get("tl_display_name") or tl_name).strip()

    if not tl_name or not tl_pin:
        return jsonify({"ok": False, "reason": "missing"}), 400

    db = SessionLocal()
    try:
        name_norm = _norm(tl_name)

        # Check if this is a manager (super user)
        managers = ["jawad", "biju", "hossam"]
        is_manager_name = name_norm in managers

        if is_manager_name and tl_pin == "112233":
            # Manager login with master passcode
            session.permanent = True
            set_tl_session(tl_name, tl_display_name)
            session['is_manager'] = True

            # Create or update TL user with manager role
            tl_user = db.query(TLUser).filter(TLUser.name_norm == name_norm).first()
            if not tl_user:
                tl_user = TLUser(
                    name_norm=name_norm,
                    display_name=tl_display_name,
                    pin_hash=generate_password_hash(tl_pin),
                    role='manager'
                )
                db.add(tl_user)
            else:
                tl_user.role = 'manager'
            db.commit()

            return jsonify({"ok": True, "manager": True})

        # Look up TL user
        tl_user = db.query(TLUser).filter(TLUser.name_norm == name_norm).first()

        if not tl_user:
            # First time - create TL user
            tl_user = TLUser(
                name_norm=name_norm,
                display_name=tl_display_name,
                pin_hash=generate_password_hash(tl_pin)
            )
            db.add(tl_user)
            db.commit()

            session.permanent = True
            set_tl_session(tl_name, tl_display_name)
            return jsonify({"ok": True, "created": True})
        else:
            # Existing user - check PIN
            if not tl_user.pin_hash:
                # Set PIN if not set
                tl_user.pin_hash = generate_password_hash(tl_pin)
                db.commit()

                session.permanent = True
                set_tl_session(tl_name, tl_display_name)
                return jsonify({"ok": True, "pin_set": True})
            elif not check_password_hash(tl_user.pin_hash, tl_pin):
                return jsonify({"ok": False, "reason": "bad_pin"}), 403
            else:
                # Successful login
                session.permanent = True
                set_tl_session(tl_name, tl_display_name)
                return jsonify({"ok": True})
    finally:
        db.close()

@app.route('/api/reconcile/state')
def api_reconcile_state():
    """Get reconciliation state for TL"""
    if not require_tl():
        return jsonify({"ok": False, "reason": "unauthorized"}), 401

    loc = (request.args.get("location") or "").strip()
    wh = (request.args.get("warehouse") or "").strip()
    line_code = (request.args.get("line_code") or "").strip()

    db = SessionLocal()
    try:
        line = db.query(Line).filter_by(location=loc, warehouse=wh, line_code=line_code).first()
        if not line:
            return jsonify({"ok": False, "reason": "not_configured"}), 404

        job = db.query(ScanJob).filter_by(line_id=line.id, status="open").order_by(ScanJob.id.desc()).first()
        if not job:
            job = ScanJob(line_id=line.id, status="open", opened_at=abu_dhabi_now())
            db.add(job)
            db.commit()

        scanned_total = db.query(func.coalesce(func.sum(Scan.qty), 0)).filter_by(job_id=job.id).scalar() or 0
        asg = db.query(Assignment).filter_by(line_id=line.id).order_by(Assignment.id.desc()).first()
        assigned = [asg.counter_name_1 if asg else "", asg.counter_name_2 if asg else ""]

        return jsonify({
            "ok": True,
            "line_id": line.id,
            "job_id": job.id,
            "target_qty": int(line.target_qty or 0),
            "scanned_total": int(scanned_total),
            "assigned": assigned,
            "status": job.status
        })
    finally:
        db.close()

@app.route('/api/reconcile/edit_target', methods=['POST'])
def api_reconcile_edit_target():
    """Edit target quantity for TL"""
    @after_this_request
    def _nc(r):
        return _no_cache(r)

    if not require_tl():
        return jsonify({"ok": False, "reason": "unauthorized"}), 401

    d = request.get_json(force=True)
    line_id = int(d.get("line_id") or 0)
    new_target = int(d.get("new_target") or 0)

    if not line_id or new_target < 0:
        return jsonify({"ok": False, "reason": "bad_input"}), 400

    db = SessionLocal()
    try:
        line = db.get(Line, line_id)
        if not line:
            return jsonify({"ok": False, "reason": "not_found"}), 404

        prev = int(line.target_qty or 0)
        if new_target == prev:
            return jsonify({"ok": False, "reason": "same_target"}), 400

        line.target_qty = new_target
        line.updated_at = abu_dhabi_now()
        db.add(line)

        # Get the open job for logging
        job = db.query(ScanJob).filter_by(line_id=line_id, status="open").order_by(ScanJob.id.desc()).first()
        if job:
            tl_session = session.get(SESSION_TL_KEY, {})
            tl_name = tl_session.get("display_name", "TL")
            rec = Reconciliation(
                job_id=job.id,
                requested_by=tl_name,  # Add required field
                reason=f"Target changed from {prev} to {new_target}",
                previous_target=prev,
                new_target=new_target,
                result="edited_target",
                approved_at=abu_dhabi_now(),
                tl_approved_by=tl_name
            )
            db.add(rec)

        # Add audit log
        tl_session = session.get(SESSION_TL_KEY, {})
        audit = AuditLog(
            actor=tl_session.get("display_name", "TL"),
            action='TARGET_UPDATED',
            entity='LINE',
            entity_id=line_id,
            payload_json=json.dumps({"previous_target": prev, "new_target": new_target})
        )
        db.add(audit)

        db.commit()
        return jsonify({
            "ok": True,
            "target_qty": new_target,
            "previous_target": prev,
            "message": "Target updated successfully"
        })
    finally:
        db.close()

@app.route('/api/reconcile/approve_variance', methods=['POST'])
def api_reconcile_approve_variance():
    """Approve variance for TL"""
    @after_this_request
    def _nc(r):
        return _no_cache(r)

    if not require_tl():
        return jsonify({"ok": False, "reason": "unauthorized"}), 401

    d = request.get_json(force=True)
    job_id = int(d.get("job_id") or 0)
    note = (d.get("note") or "").strip()

    db = SessionLocal()
    try:
        job = db.get(ScanJob, job_id)
        if not job:
            return jsonify({"ok": False, "reason": "not_found"}), 404

        job.status = "variance_approved"
        db.add(job)

        tl_session = session.get(SESSION_TL_KEY, {})
        tl_name = tl_session.get("display_name", "TL")
        rec = Reconciliation(
            job_id=job_id,
            requested_by=tl_name,  # Add the required requested_by field
            reason=note,
            result="approved_variance",
            approved_at=abu_dhabi_now(),
            tl_approved_by=tl_name
        )
        db.add(rec)
        db.commit()
        return jsonify({"ok": True, "status": job.status})
    finally:
        db.close()

@app.route('/api/lines')
def api_lines():
    """
    Query params: location, warehouse
    Returns configured lines for this LW with target & assigned counters.
    For TL users, only return lines they manage.
    """
    @after_this_request
    def _nc(r):
        return _no_cache(r)

    location = request.args.get('location', '').strip()
    warehouse = request.args.get('warehouse', '').strip()

    if not location or not warehouse:
        return jsonify({'ok': False, 'reason': 'missing'}), 400

    db = SessionLocal()
    try:
        lines_query = db.query(Line).filter(
            Line.location == location,
            Line.warehouse == warehouse
        )

        # If TL is authenticated, filter by their lines only (except for managers)
        tl_session = session.get(SESSION_TL_KEY)
        is_manager = session.get('is_manager', False)

        if tl_session and not is_manager:
            tl_name = tl_session.get("display_name", "")
            print(f"DEBUG: TL session found - {tl_name}, filtering lines")

            # Get lines where this TL is assigned (case-insensitive comparison)
            tl_line_ids = db.query(Assignment.line_id).filter(
                func.lower(Assignment.tl_name) == func.lower(tl_name),
                Assignment.active == True
            ).subquery()
            lines_query = lines_query.filter(Line.id.in_(tl_line_ids))
        elif tl_session and is_manager:
            print(f"DEBUG: Manager session found - showing all lines")
        else:
            print(f"DEBUG: No TL session found - showing no lines for regular users")
            # For regular users without TL session, return empty list
            lines_query = lines_query.filter(Line.id == -1)  # This will return no results

        lines = lines_query.all()
        print(f"DEBUG: Found {len(lines)} lines for location={location}, warehouse={warehouse}")

        out = []
        for line in lines:
            assignment = db.query(Assignment).filter(
                Assignment.line_id == line.id,
                Assignment.active == True
            ).first()

            assigned = []
            tl_name_for_line = "Unknown"
            if assignment:
                # Filter out empty or None values and include both counters
                assigned = [name for name in [assignment.counter_name_1, assignment.counter_name_2] if name and name.strip()]
                tl_name_for_line = assignment.tl_name

            print(f"DEBUG: Line {line.line_code} managed by {tl_name_for_line}, assigned to: {assigned}")

            out.append({
                'line_id': line.id,
                'line_code': line.line_code,
                'target_qty': line.target_qty,
                'assigned': assigned
            })

        return jsonify({'ok': True, 'lines': out})

    finally:
        db.close()

@app.route('/api/line/upsert', methods=['POST'])
def api_line_upsert():
    """Create/update line with target and counter assignments"""

    data = request.get_json()
    location = data.get('location')
    warehouse = data.get('warehouse')
    line_code = data.get('line_code')
    target_qty = data.get('target_qty')
    counter1 = data.get('counter1')
    counter2 = data.get('counter2')
    tl_name = data.get('tl_name')
    pin = data.get('pin')

    if not all([location, warehouse, line_code, target_qty, counter1, counter2, tl_name, pin]):
        return jsonify({'error': 'Missing required fields'}), 400

    db = SessionLocal()
    try:
        # Find or create line
        line = db.query(Line).filter(
            Line.location == location,
            Line.warehouse == warehouse,
            Line.line_code == line_code
        ).first()

        _, tl_norm = _session_user()

        if line:
            line.target_qty = target_qty
            line.updated_at = abu_dhabi_now()
        else:
            line = Line(
                location=location,
                warehouse=warehouse,
                line_code=line_code,
                target_qty=target_qty,
                created_by_tl_norm=tl_norm
            )
            db.add(line)
            db.flush()

        # Deactivate old assignments
        db.query(Assignment).filter(Assignment.line_id == line.id).update({'active': False})

        # Create new assignment
        assignment = Assignment(
            line_id=line.id,
            counter_name_1=counter1,
            counter_name_2=counter2,
            tl_name=tl_name,
            tl_pin_hash=hash_pin(pin),
            active=True
        )
        db.add(assignment)

        # Ensure one open job exists
        open_job = db.query(ScanJob).filter(
            ScanJob.line_id == line.id,
            ScanJob.status == 'open'
        ).first()

        if not open_job:
            job = ScanJob(
                line_id=line.id,
                status='open',
                opened_by=tl_name,
                opened_at=abu_dhabi_now() # Updated here
            )
            db.add(job)

        # Audit log
        audit = AuditLog(
            actor=tl_name,
            action='LINE_SETUP',
            entity='LINE',
            entity_id=line.id,
            payload_json=json.dumps(data)
        )
        db.add(audit)

        db.commit()

        # Return fresh state
        return jsonify({
            'ok': True,
            'line_id': line.id,
            'line_code': line.line_code,
            'target_qty': line.target_qty,
            'assigned': [counter1, counter2]
        })

    except Exception as e:
        db.rollback()
        return jsonify({'error': str(e)}), 500
    finally:
        db.close()

@app.route('/api/job/state')
def api_job_state():
    """Get current job state for a line"""
    location = request.args.get('location', '').strip()
    warehouse = request.args.get('warehouse', '').strip()
    line_code = request.args.get('line_code', '').strip()
    counter = request.args.get('counter', '').strip()

    if not all([location, warehouse, line_code]):
        return jsonify({'ok': False, 'reason': 'missing_params'}), 400

    db = SessionLocal()
    try:
        line = db.query(Line).filter(
            Line.location == location,
            Line.warehouse == warehouse,
            Line.line_code == line_code
        ).first()

        if not line:
            return jsonify({'ok': False, 'reason': 'not_configured'}), 404

        assignment = db.query(Assignment).filter(
            Assignment.line_id == line.id,
            Assignment.active == True
        ).first()

        assigned = []
        is_assigned = False
        if assignment:
            assigned = [assignment.counter_name_1, assignment.counter_name_2]
            # Normalize counter names for comparison (case-insensitive)
            counter_norm = _norm(counter)
            is_assigned = counter_norm in {_norm(name) for name in assigned if name}

        # Get current job (including locked ones)
        current_job = db.query(ScanJob).filter(
            ScanJob.line_id == line.id,
            ScanJob.status.in_(['open', 'locked_recon', 'variance_approved'])
        ).first()

        # Check if there are completed jobs
        completed_jobs = db.query(ScanJob).filter(
            ScanJob.line_id == line.id,
            ScanJob.status == 'submitted'
        ).count()

        # If there are completed jobs and no current jobs, the line is completed
        if completed_jobs > 0 and not current_job:
            return jsonify({
                'ok': False,
                'reason': 'line_completed',
                'message': 'This line has been completed. Contact your Team Leader for a new assignment.'
            }), 410

        # If no current job exists and no completed jobs, create a new open job
        if not current_job:
            current_job = ScanJob(
                line_id=line.id,
                status='open',
                opened_at=abu_dhabi_now()
            )
            db.add(current_job)
            db.commit()

        # Count scanned total using sum of qty
        scanned_total = db.query(func.coalesce(func.sum(Scan.qty), 0)).filter(Scan.job_id == current_job.id).scalar() or 0

        return jsonify({
            'ok': True,
            'line_id': line.id,
            'job_id': current_job.id,
            'target': int(line.target_qty or 0),
            'target_qty': int(line.target_qty or 0),
            'assigned': assigned,
            'is_assigned': bool(is_assigned),
            'scanned_total': int(scanned_total),
            'status': current_job.status
        })

    finally:
        db.close()

def _ns(s):
    """Normalize string for scan comparison"""
    import re
    return re.sub(r"[^A-Za-z0-9\-_]", "", (s or "").strip()).upper()

@app.route('/api/scan/add', methods=['POST'])
def api_scan_add():
    """Add a scan to the job with strict duplicate checking"""
    data = request.get_json(force=True)
    job_id = int(data.get("job_id") or 0)
    line_id = int(data.get("line_id") or 0)
    counter_name = (data.get("counter_name") or "").strip()
    sku_raw = (data.get("sku") or "").strip()
    code_raw = (data.get("serial_or_code") or "").strip()
    qty = int(data.get("qty") or 1)
    source = (data.get("source") or "manual").strip()

    if not (job_id and line_id and counter_name and code_raw and qty >= 1):
        return jsonify({"ok": False, "reason": "missing"}), 400

    sku = _ns(sku_raw)
    code = _ns(code_raw)

    db = SessionLocal()
    try:
        # Strict duplicate check: same (job_id, sku, serial_code) combination only
        existing = db.query(Scan.id).filter(
            Scan.job_id == job_id,
            Scan.sku == sku,
            Scan.serial_code == code
        ).first()

        if existing:
            return jsonify({"ok": False, "duplicate": True}), 409

        # Add scan
        scan = Scan(
            job_id=job_id,
            line_id=line_id,
            counter_name=counter_name,
            sku=sku,
            serial_code=code,
            qty=qty,
            source=source,
            created_at=abu_dhabi_now()
        )
        db.add(scan)
        db.commit()

        # Get updated total count
        scanned_total = db.query(func.coalesce(func.sum(Scan.qty), 0)).filter(Scan.job_id == job_id).scalar() or 0

        return jsonify({"ok": True, "scanned_total": int(scanned_total)})

    except Exception as e:
        db.rollback()
        return jsonify({'ok': False, 'error': 'Failed to add item'}), 500
    finally:
        db.close()

@app.route('/api/submit/final', methods=['POST'])
def api_submit_final():
    """Finalize and submit job - only when variance approved or total matches target"""
    data = request.get_json()
    job_id = data.get('job_id')
    counter_name = data.get('counter_name')

    db = SessionLocal()
    try:
        job = db.query(ScanJob).filter(ScanJob.id == job_id).first()
        if not job:
            return jsonify({'ok': False, 'error': 'Job not found'}), 404

        line = job.line
        scanned_total = db.query(func.coalesce(func.sum(Scan.qty), 0)).filter(Scan.job_id == job_id).scalar() or 0

        # Check if submission allowed
        allow = (job.status == "variance_approved") or (scanned_total == int(line.target_qty or 0))
        if not allow:
            return jsonify({"ok": False, "reason": "mismatch"}), 412

        # Get all scans for export
        scans = db.query(Scan).filter(Scan.job_id == job_id).all()

        # Export to Excel
        with FileLock(LOCK_PATH, timeout=10):
            wb = load_workbook(MDF_PATH)
            ws = wb.active

            for scan in scans:
                now = scan.created_at
                row = [
                    now.strftime("%Y-%m-%d"),  # Date
                    now.strftime("%H:%M:%S"),  # Time
                    line.location,             # Location
                    line.warehouse,            # Warehouse
                    scan.counter_name,         # CounterName
                    scan.sku or '',            # SKU
                    scan.serial_code,          # SerialOrCode
                    scan.qty,                  # QTY
                    scan.source                # Source
                ]
                ws.append(row)

            wb.save(MDF_PATH)
            wb.close()

        # Close job
        job.status = 'submitted'
        job.closed_at = abu_dhabi_now()

        # Audit log
        audit = AuditLog(
            actor=counter_name or 'Unknown',
            action='JOB_SUBMIT',
            entity='SCANJOB',
            entity_id=job_id,
            payload_json=json.dumps({'scanned_total': scanned_total, 'target': line.target_qty})
        )
        db.add(audit)

        db.commit()

        return jsonify({'ok': True, 'submitted': True})

    except Exception as e:
        db.rollback()
        return jsonify({'ok': False, 'error': str(e)}), 500
    finally:
        db.close()

@app.route('/api/reconcile/request', methods=['POST'])
def api_reconcile_request():
    """Request reconciliation and lock job until TL acts"""
    data = request.get_json(force=True)
    job_id = int(data.get("job_id") or 0)
    line_id = int(data.get("line_id") or 0)
    counter_name = (data.get("counter_name") or "").strip()
    reason = (data.get("reason") or "").strip()

    if not (job_id and line_id and counter_name):
        return jsonify({"ok": False, "reason": "missing"}), 400

    db = SessionLocal()
    try:
        job = db.get(ScanJob, job_id)
        line = db.get(Line, line_id)
        if not job or not line:
            return jsonify({"ok": False, "reason": "not_found"}), 404

        # Get current scanned total
        scanned_total = db.query(func.coalesce(func.sum(Scan.qty), 0)).filter_by(job_id=job_id).scalar() or 0
        target = int(line.target_qty or 0)

        # Check if reconciliation needed
        if scanned_total == target:
            return jsonify({"ok": False, "reason": "no_mismatch"}), 400

        # Get TL name from assignment
        asg = db.query(Assignment).filter_by(line_id=line_id, active=True).order_by(Assignment.id.desc()).first()
        tl_name_norm = _norm(asg.tl_name) if asg and getattr(asg, "tl_name", None) else ""

        # Create reconciliation request
        req = ReconciliationRequest(
            line_id=line_id,
            job_id=job_id,
            tl_name_norm=tl_name_norm,
            requested_by=counter_name,
            reason=reason,
            requested_qty=int(scanned_total),
            status='pending'
        )

        # Lock job until TL acts
        job.status = "locked_recon"

        db.add(req)
        db.add(job)
        db.commit()

        return jsonify({"ok": True, "requested_qty": int(scanned_total)})

    except Exception as e:
        db.rollback()
        return jsonify({"ok": False, "error": str(e)}), 500
    finally:
        db.close()

@app.route('/api/reconcile/approve', methods=['POST'])
def api_reconcile_approve():
    """TL approve reconciliation"""
    if not session.get('tl_authenticated'):
        return jsonify({'error': 'TL authentication required'}), 401

    data = request.get_json()
    job_id = data.get('job_id')
    mode = data.get('mode')  # edit_target or approve_variance
    new_target = data.get('new_target')
    note = data.get('note')

    db = SessionLocal()
    try:
        job = db.query(ScanJob).filter(ScanJob.id == job_id).first()
        if not job:
            return jsonify({'error': 'Job not found'}), 404

        reconciliation = db.query(Reconciliation).filter(
            Reconciliation.job_id == job_id
        ).order_by(Reconciliation.id.desc()).first()

        if mode == 'edit_target' and new_target:
            job.line.target_qty = new_target
            job.line.updated_at = abu_dhabi_now()
            reconciliation.result = 'edited_target'
            reconciliation.new_target = new_target
        elif mode == 'approve_variance':
            job.status = 'variance_approved'
            reconciliation.result = 'approved_variance'

        reconciliation.tl_approved_by = session.get('tl_name') # This should also be updated to display_name if used
        reconciliation.approved_at = abu_dhabi_now()
        reconciliation.note = note

        db.commit()
        return jsonify({'success': True})

    except Exception as e:
        db.rollback()
        return jsonify({'error': str(e)}), 500
    finally:
        db.close()

@app.route('/api/recent-scans')
def api_recent_scans():
    """Get recent scans for a job"""
    job_id = request.args.get('job_id')
    if not job_id:
        return jsonify([])

    db = SessionLocal()
    try:
        scans = db.query(Scan).filter(
            Scan.job_id == job_id
        ).order_by(Scan.created_at.desc()).limit(10).all()

        scan_data = []
        for scan in scans:
            scan_data.append({
                'sku': scan.sku or '',
                'serial_code': scan.serial_code,
                'qty': scan.qty,
                'counter_name': scan.counter_name,
                'source': scan.source,
                'time': scan.created_at.strftime('%H:%M:%S')
            })

        return jsonify(scan_data)

    finally:
        db.close()

@app.route('/exports/MDF.xlsx')
def download_excel():
    """Download the Excel file with all completed job data"""
    db = SessionLocal()
    try:
        # Get only submitted jobs from database (not variance_approved)
        jobs = db.query(ScanJob).filter(
            ScanJob.status == 'submitted'
        ).all()

        # Regenerate Excel file with current data
        with FileLock(LOCK_PATH, timeout=10):
            wb = Workbook()
            ws = wb.active
            ws.title = "Sheet1"
            ws.append(COLUMNS)

            # Add all scans from completed jobs
            for job in jobs:
                scans = db.query(Scan).filter(Scan.job_id == job.id).all()
                for scan in scans:
                    now = scan.created_at # This will be in Abu Dhabi time
                    row = [
                        now.strftime("%Y-%m-%d"),  # Date
                        now.strftime("%H:%M:%S"),  # Time
                        job.line.location,         # Location
                        job.line.warehouse,        # Warehouse
                        scan.counter_name,         # CounterName
                        scan.sku or '',            # SKU
                        scan.serial_code,          # SerialOrCode
                        scan.qty,                  # QTY
                        scan.source                # Source
                    ]
                    ws.append(row)

            wb.save(MDF_PATH)
            wb.close()

        return send_file(MDF_PATH, as_attachment=True, download_name='MDF.xlsx')
    finally:
        db.close()

@app.route('/api/reconcile/tl_queue')
def api_tl_reconcile_queue():
    """Get pending reconciliation requests for TL"""
    if not require_tl():
        return jsonify({"ok": False, "reason": "unauthorized"}), 401

    tl_session = session.get(SESSION_TL_KEY, {})
    tl_name = tl_session.get("display_name", "")

    db = SessionLocal()
    try:
        # Get TL's lines using case-insensitive comparison
        assignments = db.query(Assignment).filter(
            func.lower(Assignment.tl_name) == func.lower(tl_name),
            Assignment.active == True
        ).all()

        line_ids = [a.line_id for a in assignments]

        if not line_ids:
            return jsonify({"ok": True, "requests": []})

        # Get pending reconciliation requests for TL's lines
        requests = db.query(ReconciliationQueue).filter(
            ReconciliationQueue.line_id.in_(line_ids),
            ReconciliationQueue.status == 'pending'
        ).order_by(ReconciliationQueue.created_at.desc()).all()

        queue_data = []
        for req in requests:
            queue_data.append({
                'id': req.id,
                'job_id': req.job_id,
                'line_code': req.line.line_code,
                'location': req.line.location,
                'warehouse': req.line.warehouse,
                'requested_by': req.requested_by,
                'reason': req.reason,
                'scanned_total': req.scanned_total,
                'target_qty': req.target_qty,
                'created_at': req.created_at.strftime('%H:%M:%S')
            })

        return jsonify({"ok": True, "requests": queue_data})

    finally:
        db.close()

@app.route('/api/reconcile/notification_count')
def api_reconcile_notification_count():
    """Get count of pending reconciliation requests for TL"""
    if not require_tl():
        return jsonify({"ok": False, "reason": "unauthorized"}), 401

    tl_session = session.get(SESSION_TL_KEY, {})
    tl_name = tl_session.get("display_name", "")

    db = SessionLocal()
    try:
        # Get TL's lines using case-insensitive comparison
        assignments = db.query(Assignment).filter(
            func.lower(Assignment.tl_name) == func.lower(tl_name),
            Assignment.active == True
        ).all()

        line_ids = [a.line_id for a in assignments]

        if not line_ids:
            return jsonify({"ok": True, "count": 0})

        # Count pending reconciliation requests for TL's lines
        count = db.query(ReconciliationQueue).filter(
            ReconciliationQueue.line_id.in_(line_ids),
            ReconciliationQueue.status == 'pending'
        ).count()

        return jsonify({"ok": True, "count": count})

    finally:
        db.close()

@app.route('/api/reconcile/pending_count_all')
def api_reconcile_pending_count_all():
    """Get count of all pending reconciliation requests (no TL auth required)"""
    db = SessionLocal()
    try:
        # Count all pending reconciliation requests
        count = db.query(ReconciliationQueue).filter(
            ReconciliationQueue.status == 'pending'
        ).count()

        return jsonify({"ok": True, "count": count})

    finally:
        db.close()



@app.route('/api/reconcile/tl_respond', methods=['POST'])
def api_tl_respond_reconcile():
    """TL responds to reconciliation request"""
    if not require_tl():
        return jsonify({"ok": False, "reason": "unauthorized"}), 401

    data = request.get_json()
    queue_id = data.get('queue_id')
    action = data.get('action')  # 'approve_variance' or 'edit_target'
    new_target = data.get('new_target')
    note = data.get('note', '')

    db = SessionLocal()
    try:
        queue_item = db.query(ReconciliationQueue).filter(ReconciliationQueue.id == queue_id).first()
        if not queue_item:
            return jsonify({'error': 'Request not found'}), 404

        tl_session = session.get(SESSION_TL_KEY, {})
        tl_name = tl_session.get("display_name", "TL")

        if action == 'edit_target' and new_target:
            # Update target quantity
            line = queue_item.line
            old_target = line.target_qty
            line.target_qty = new_target
            line.updated_at = abu_dhabi_now()

            queue_item.status = 'approved'
            queue_item.tl_response = f"Target updated from {old_target} to {new_target}. {note}"
            queue_item.resolved_at = abu_dhabi_now()

        elif action == 'approve_variance':
            # Approve variance - update job status
            job = queue_item.job
            job.status = 'variance_approved'

            queue_item.status = 'approved'
            queue_item.tl_response = f"Variance approved. {note}"
            queue_item.resolved_at = abu_dhabi_now()

        # Add reconciliation record
        reconciliation = Reconciliation(
            job_id=queue_item.job_id,
            requested_by=queue_item.requested_by,
            reason=queue_item.reason,
            previous_target=queue_item.target_qty,
            new_target=new_target if action == 'edit_target' else queue_item.target_qty,
            result=action,
            approved_at=abu_dhabi_now(),
            tl_approved_by=tl_name,
            note=note
        )
        db.add(reconciliation)

        db.commit()
        return jsonify({'success': True})

    except Exception as e:
        db.rollback()
        return jsonify({'error': str(e)}), 500
    finally:
        db.close()

@app.route('/api/reconcile/check_response')
def api_check_reconcile_response():
    """Check if counter's reconciliation request has been responded to"""
    job_id = request.args.get('job_id')
    if not job_id:
        return jsonify({'error': 'Missing job_id'}), 400

    db = SessionLocal()
    try:
        # Check if there's a resolved request for this job that hasn't been acknowledged yet
        resolved_request = db.query(ReconciliationQueue).filter(
            ReconciliationQueue.job_id == job_id,
            ReconciliationQueue.status == 'approved',
            ReconciliationQueue.acknowledged == False
        ).order_by(ReconciliationQueue.resolved_at.desc()).first()

        if resolved_request:
            return jsonify({
                'resolved': True,
                'response': resolved_request.tl_response,
                'resolved_at': resolved_request.resolved_at.strftime('%H:%M:%S'),
                'acknowledged': False
            })
        else:
            return jsonify({'resolved': False})

    finally:
        db.close()

@app.route('/api/reconcile/acknowledge', methods=['POST'])
def api_reconcile_acknowledge():
    """Acknowledge TL response to reconciliation request"""
    data = request.get_json()
    job_id = data.get('job_id')

    if not job_id:
        return jsonify({'error': 'Missing job_id'}), 400

    db = SessionLocal()
    try:
        # Find the resolved request and mark it as acknowledged
        resolved_request = db.query(ReconciliationQueue).filter(
            ReconciliationQueue.job_id == job_id,
            ReconciliationQueue.status == 'approved',
            ReconciliationQueue.acknowledged == False
        ).first()

        if resolved_request:
            resolved_request.acknowledged = True
            db.commit()
            return jsonify({'success': True, 'acknowledged': True})
        else:
            return jsonify({'success': True, 'message': 'No pending response to acknowledge'})

    except Exception as e:
        db.rollback()
        return jsonify({'error': str(e)}), 500
    finally:
        db.close()

@app.route('/api/logs/delete/<int:job_index>', methods=['DELETE'])
def api_delete_log(job_index):
    """Delete a specific log entry (database or historical)"""
    if not require_tl():
        return jsonify({"error": "TL authentication required"}), 401

    data = request.get_json() or {}
    passcode = data.get('passcode', '')

    # Verify passcode for historical log deletion
    if passcode != '240986':
        return jsonify({"error": "Invalid passcode"}), 403

    db = SessionLocal()
    try:
        # Get all completed jobs from database
        db_jobs = db.query(ScanJob).filter(
            ScanJob.status.in_(['submitted', 'variance_approved'])
        ).order_by(ScanJob.closed_at.desc()).all()

        # Get historical jobs from Excel
        historical_jobs = []
        try:
            if os.path.exists(MDF_PATH):
                df = pd.read_excel(MDF_PATH)
                if not df.empty and len(df) > 0:
                    # Group by date, location, warehouse, and first counter name to identify unique jobs
                    historical_groups = df.groupby(['Date', 'Location', 'Warehouse', 'CounterName']).agg({
                        'QTY': 'sum',
                        'SerialOrCode': 'count',
                        'Time': 'max'
                    }).reset_index()

                    for _, row in historical_groups.iterrows():
                        try:
                            date_str = str(row['Date'])
                            time_str = str(row['Time'])
                            if 'T' in date_str:
                                closed_at = datetime.fromisoformat(date_str.replace('Z', '+00:00'))
                            else:
                                datetime_str = f"{date_str} {time_str}"
                                closed_at = datetime.strptime(datetime_str, '%Y-%m-%d %H:%M:%S')
                        except:
                            closed_at = abu_dhabi_now()

                        historical_jobs.append({
                            'date': str(row['Date']),
                            'location': row['Location'],
                            'warehouse': row['Warehouse'],
                            'counter': row['CounterName'],
                            'closed_at': closed_at,
                            'source': 'excel'
                        })
        except Exception as e:
            print(f"Error reading historical data: {e}")

        # Sort all jobs by closed_at (newest first)
        all_jobs = []

        # Add database jobs
        for job in db_jobs:
            all_jobs.append({
                'type': 'database',
                'job': job,
                'closed_at': job.closed_at,
                'source': 'database'
            })

        # Add historical jobs
        for hist_job in historical_jobs:
            all_jobs.append({
                'type': 'historical',
                'job': hist_job,
                'closed_at': hist_job['closed_at'],
                'source': 'excel'
            })

        # Sort by closed_at (newest first)
        all_jobs.sort(key=lambda x: x['closed_at'], reverse=True)

        if job_index >= len(all_jobs):
            return jsonify({"error": "Job not found"}), 404

        job_to_delete = all_jobs[job_index]

        if job_to_delete['type'] == 'database':
            # Delete database job
            job = job_to_delete['job']

            # Delete related scans first
            db.query(Scan).filter(Scan.job_id == job.id).delete()
            # Delete related reconciliations
            db.query(Reconciliation).filter(Reconciliation.job_id == job.id).delete()
            # Delete related reconciliation queue items
            db.query(ReconciliationQueue).filter(ReconciliationQueue.job_id == job.id).delete()
            # Delete the job
            db.delete(job)

            # Add audit log
            tl_session = session.get(SESSION_TL_KEY, {})
            audit = AuditLog(
                actor=tl_session.get("display_name", "TL"),
                action='LOG_DELETE',
                entity='SCANJOB',
                entity_id=job.id,
                payload_json=json.dumps({"line_code": job.line.line_code, "type": "database"})
            )
            db.add(audit)
            db.commit()

        elif job_to_delete['type'] == 'historical':
            # Delete historical job from Excel
            hist_job = job_to_delete['job']

            with FileLock(LOCK_PATH, timeout=10):
                df = pd.read_excel(MDF_PATH)

                # Remove rows matching this historical job
                mask = (
                    (df['Date'].astype(str) == hist_job['date']) &
                    (df['Location'] == hist_job['location']) &
                    (df['Warehouse'] == hist_job['warehouse']) &
                    (df['CounterName'] == hist_job['counter'])
                )

                df_filtered = df[~mask]

                # Save back to Excel
                from openpyxl import Workbook
                wb = Workbook()
                ws = wb.active
                ws.title = "Sheet1"

                # Add headers
                ws.append(COLUMNS)

                # Add remaining data
                for _, row in df_filtered.iterrows():
                    ws.append([
                        row['Date'], row['Time'], row['Location'], row['Warehouse'],
                        row['CounterName'], row['SKU'], row['SerialOrCode'],
                        row['QTY'], row['Source']
                    ])

                wb.save(MDF_PATH)
                wb.close()

            # Add audit log
            tl_session = session.get(SESSION_TL_KEY, {})
            audit = AuditLog(
                actor=tl_session.get("display_name", "TL"),
                action='HISTORICAL_LOG_DELETE',
                entity='EXCEL',
                payload_json=json.dumps({
                    "date": hist_job['date'],
                    "location": hist_job['location'],
                    "warehouse": hist_job['warehouse'],
                    "counter": hist_job['counter'],
                    "type": "historical"
                })
            )
            db.add(audit)
            db.commit()

        return jsonify({"success": True})

    except Exception as e:
        db.rollback()
        return jsonify({"error": str(e)}), 500
    finally:
        db.close()

@app.route('/api/job/reset', methods=['POST'])
def api_job_reset():
    """Reset job to open status with passcode"""
    data = request.get_json(force=True)
    job_id = int(data.get("job_id") or 0)
    passcode = (data.get("passcode") or "").strip()

    if passcode != "240986":
        return jsonify({"ok": False, "reason": "bad_passcode"}), 403

    db = SessionLocal()
    try:
        job = db.get(ScanJob, job_id)
        if not job:
            return jsonify({"ok": False, "reason": "not_found"}), 404

        job.status = "open"
        db.add(job)
        db.commit()

        return jsonify({"ok": True})

    except Exception as e:
        db.rollback()
        return jsonify({"ok": False, "error": str(e)}), 500
    finally:
        db.close()

@app.route('/api/lines/inline_reconcile', methods=['POST'])
def api_lines_inline_reconcile():
    """Inline approve/edit from Line Management"""
    if not require_tl():
        return jsonify({"ok": False, "reason": "unauthorized"}), 401

    data = request.get_json(force=True)
    req_id = int(data.get("request_id") or 0)
    new_target = data.get("new_target", None)

    db = SessionLocal()
    try:
        req = db.get(ReconciliationRequest, req_id)
        if not req or req.status != "pending":
            return jsonify({"ok": False, "reason": "not_found_or_done"}), 404

        job = db.get(ScanJob, req.job_id)
        line = db.get(Line, req.line_id)
        if not job or not line:
            return jsonify({"ok": False, "reason": "not_found"}), 404

        # Set target based on action
        tgt = int(new_target) if new_target is not None else int(req.requested_qty or 0)
        line.target_qty = tgt
        line.updated_at = abu_dhabi_now()

        # Unlock job and allow submit
        job.status = "variance_approved"
        req.status = "approved"
        req.resolved_at = abu_dhabi_now()

        tl_session = session.get(SESSION_TL_KEY, {})
        req.resolved_by = tl_session.get("display_name", "TL")

        db.add(line)
        db.add(job)
        db.add(req)
        db.commit()

        return jsonify({"ok": True, "target_qty": tgt, "job_status": job.status})

    except Exception as e:
        db.rollback()
        return jsonify({"ok": False, "error": str(e)}), 500
    finally:
        db.close()

@app.route('/api/line/reset', methods=['POST'])
def api_line_reset():
    """Reset a completed line to create a new job (TL only)"""
    if not _require_tl():
        return jsonify({"ok": False, "reason": "unauthorized"}), 401

    data = request.get_json()
    line_id = data.get('line_id')
    passcode = data.get('passcode')

    if not all([line_id, passcode]):
        return jsonify({'error': 'Missing required fields'}), 400

    # Verify passcode
    if passcode != '240986':
        return jsonify({'error': 'Invalid passcode'}), 403

    db = SessionLocal()
    try:
        line = db.get(Line, line_id)
        if not line:
            return jsonify({'error': 'Line not found'}), 404

        # Verify TL permissions
        display_name, tl_norm = _session_user()
        is_manager = _is_manager()

        if not is_manager:
            # Check if TL created this line
            if line.created_by_tl_norm != tl_norm:
                return jsonify({'error': 'You can only reset lines you created'}), 403

        # Check if there's already an open job
        open_job = db.query(ScanJob).filter(
            ScanJob.line_id == line.id,
            ScanJob.status == 'open'
        ).first()

        if open_job:
            return jsonify({'error': 'Line already has an open job'}), 400

        # Create new open job
        new_job = ScanJob(
            line_id=line.id,
            status='open',
            opened_at=abu_dhabi_now(),
            opened_by=display_name
        )
        db.add(new_job)

        # Add audit log
        audit = AuditLog(
            actor=display_name,
            action='LINE_RESET',
            entity='LINE',
            entity_id=line.id,
            payload_json=json.dumps({
                'location': line.location,
                'warehouse': line.warehouse,
                'line_code': line.line_code
            })
        )
        db.add(audit)

        db.commit()
        return jsonify({'ok': True, 'message': 'Line reset successfully. Counters can now start counting again.'})

    except Exception as e:
        db.rollback()
        return jsonify({'error': str(e)}), 500
    finally:
        db.close()

@app.route('/api/counter/jobs')
def api_counter_jobs():
    """Get jobs for a specific counter - only non-submitted jobs"""
    counter = request.args.get('counter', '').strip()

    if not counter:
        return jsonify({"ok": False, "reason": "missing_counter"}), 400

    db = SessionLocal()
    try:
        cnorm = counter.strip().casefold()

        # Get jobs that are not submitted
        rows = db.query(ScanJob, Line, Assignment).join(
            Line, Line.id == ScanJob.line_id
        ).outerjoin(
            Assignment, Assignment.line_id == Line.id
        ).filter(
            ScanJob.status.in_(["open", "locked_recon", "variance_approved"]),  # not submitted
            Assignment.active == True
        ).all()

        items = []
        for job, line, asg in rows:
            if not asg:
                continue
            a1 = (getattr(asg, "counter_name_1", "") or "").strip().casefold()
            a2 = (getattr(asg, "counter_name_2", "") or "").strip().casefold()
            if cnorm in {a1, a2}:
                items.append({
                    "job_id": job.id,
                    "line_id": line.id,
                    "location": line.location,
                    "warehouse": line.warehouse,
                    "line_code": line.line_code,
                    "target_qty": int(line.target_qty or 0),
                    "status": job.status
                })

        return jsonify({"ok": True, "items": items})

    finally:
        db.close()

@app.route('/api/reconcile/inbox')
def api_reconcile_inbox():
    """Get pending reconciliation requests for TL"""
    if not _require_tl():
        return jsonify({"ok": False, "reason": "unauthorized"}), 401

    _, tl_norm = _session_user()

    db = SessionLocal()
    try:
        # Get pending reconciliation requests for this TL
        requests = db.query(ReconciliationRequest).filter(
            ReconciliationRequest.tl_name_norm == tl_norm,
            ReconciliationRequest.status == 'pending'
        ).order_by(ReconciliationRequest.created_at.desc()).all()

        inbox_data = []
        for req in requests:
            inbox_data.append({
                'request_id': req.id,
                'job_id': req.job_id,
                'line_code': req.line.line_code,
                'location': req.line.location,
                'warehouse': req.line.warehouse,
                'requested_by': req.requested_by,
                'requested_qty': req.requested_qty,
                'target_qty': req.line.target_qty,
                'created_at': req.created_at.strftime('%H:%M:%S')
            })

        return jsonify({"ok": True, "requests": inbox_data})

    finally:
        db.close()

@app.route('/api/reconcile/resolve', methods=['POST'])
def api_reconcile_resolve():
    """TL resolves reconciliation request and unlocks job"""
    if not _require_tl():
        return jsonify({"ok": False, "reason": "unauthorized"}), 401

    data = request.get_json(force=True)
    request_id = int(data.get("request_id") or 0)
    action = data.get("action", "approve_variance")
    new_target = data.get("new_target", None)

    db = SessionLocal()
    try:
        req = db.get(ReconciliationRequest, request_id)
        if not req or req.status != "pending":
            return jsonify({"ok": False, "reason": "not_found_or_done"}), 404

        job = db.get(ScanJob, req.job_id)
        line = db.get(Line, req.line_id)
        if not job or not line:
            return jsonify({"ok": False, "reason": "not_found"}), 404

        # Set target based on action
        if action == "edit_target" and new_target is not None:
            line.target_qty = int(new_target)
        elif req.requested_qty is not None:
            line.target_qty = req.requested_qty

        # Unlock job and allow submit
        job.status = "variance_approved"
        req.status = "approved"
        req.resolved_at = abu_dhabi_now()

        display_name, _ = _session_user()
        req.resolved_by = display_name

        db.add(job)
        db.add(line)
        db.add(req)
        db.commit()

        return jsonify({"ok": True, "target_qty": int(line.target_qty)})

    except Exception as e:
        db.rollback()
        return jsonify({"ok": False, "error": str(e)}), 500
    finally:
        db.close()

@app.route('/api/reconcile/line_requests')
def api_reconcile_line_requests():
    """Get reconciliation requests for a specific line"""
    if not _require_tl():
        return jsonify({"ok": False, "reason": "unauthorized"}), 401

    line_id = int(request.args.get("line_id") or 0)
    if not line_id:
        return jsonify({"ok": False, "reason": "missing_line_id"}), 400

    db = SessionLocal()
    try:
        requests = db.query(ReconciliationRequest).filter(
            ReconciliationRequest.line_id == line_id,
            ReconciliationRequest.status == 'pending'
        ).order_by(ReconciliationRequest.created_at.desc()).all()

        request_data = []
        for req in requests:
            request_data.append({
                'id': req.id,
                'requested_by': req.requested_by,
                'requested_qty': req.requested_qty,
                'created_at': req.created_at.isoformat(),
                'job_id': req.job_id
            })

        return jsonify({"ok": True, "requests": request_data})

    finally:
        db.close()

@app.route('/api/lines/manage')
def api_lines_manage():
    """Get all lines with edit permissions for TL/Manager view"""
    if not _require_tl():
        return jsonify({"ok": False, "reason": "unauthorized"}), 401

    db = SessionLocal()
    try:
        display_name, tl_norm = _session_user()
        is_manager = _is_manager()

        # Get all lines with assignments
        lines_with_assignments = db.query(Line, Assignment).join(
            Assignment,
            (Assignment.line_id == Line.id) & (Assignment.active == True)
        ).all()

        lines_data = []
        for line, assignment in lines_with_assignments:
            # Check current job status
            current_job = db.query(ScanJob).filter(
                ScanJob.line_id == line.id,
                ScanJob.status.in_(['open', 'locked_recon', 'variance_approved'])
            ).first()

            if current_job:
                job_status = current_job.status
            else:
                job_status = 'submitted'

            # Check edit permissions
            can_edit = is_manager or (line.created_by_tl_norm == tl_norm)

            # Check for pending reconciliation requests
            pending_reconciliation = db.query(ReconciliationRequest).filter(
                ReconciliationRequest.line_id == line.id,
                ReconciliationRequest.status == 'pending'
            ).first() is not None

            line_data = {
                'id': line.id,
                'line_code': line.line_code,
                'location': line.location,
                'warehouse': line.warehouse,
                'target_qty': line.target_qty,
                'created_at': line.created_at.strftime('%Y-%m-%d %H:%M'),
                'counter_name_1': assignment.counter_name_1 if assignment else None,
                'counter_name_2': assignment.counter_name_2 if assignment else None,
                'tl_name': assignment.tl_name if assignment else 'Not assigned',
                'can_edit': can_edit,
                'status': job_status,
                'pending_reconciliation': pending_reconciliation
            }
            lines_data.append(line_data)

        return jsonify({'ok': True, 'lines': lines_data, 'current_tl': display_name, 'is_manager': is_manager})

    finally:
        db.close()

@app.route('/api/line/delete', methods=['POST'])
def api_line_delete():
    """Delete a line configuration"""
    data = request.get_json()
    location = data.get('location')
    warehouse = data.get('warehouse')
    line_code = data.get('line_code')
    tl_name = data.get('tl_name')
    pin = data.get('pin')

    # Check if this is a manager delete request
    is_manager_delete = request.headers.get('X-Manager-Delete') == 'true'

    if not all([location, warehouse, line_code]):
        return jsonify({'error': 'Missing required fields'}), 400

    # Check TL session
    tl_session = session.get(SESSION_TL_KEY)
    is_manager = session.get('is_manager', False)

    if not tl_session:
        return jsonify({'error': 'TL authentication required'}), 401

    # For manager deletes, we need different validation
    if is_manager_delete or is_manager:
        if pin != '240986':
            return jsonify({'error': 'Invalid passcode'}), 403
        # Manager can delete any line
        pass
    else:
        # Regular TL delete - require all fields
        if not all([tl_name, pin]):
            return jsonify({'error': 'Missing TL credentials'}), 400

    db = SessionLocal()
    try:
        # Find the line
        line = db.query(Line).filter(
            Line.location == location,
            Line.warehouse == warehouse,
            Line.line_code == line_code
        ).first()

        if not line:
            return jsonify({'error': 'Line not found'}), 404

        # Verify permissions
        assignment = db.query(Assignment).filter(
            Assignment.line_id == line.id,
            Assignment.active == True
        ).first()

        if not assignment:
            return jsonify({'error': 'No assignment found for this line'}), 404

        # Permission check
        if is_manager or is_manager_delete:
            # Managers can delete any line with correct passcode (already verified above)
            pass
        else:
            # TL must match and have correct PIN
            if not verify_pin(pin, assignment.tl_pin_hash):
                return jsonify({'error': 'Unauthorized or wrong PIN'}), 403

            # Additional check: ensure the requesting TL is the one assigned to this line
            current_tl_name = tl_session.get('display_name', tl_name)
            if _norm(current_tl_name) != _norm(assignment.tl_name):
                return jsonify({'error': 'You can only delete lines you created'}), 403

        # Delete related data in correct order
        # Delete scans first
        jobs = db.query(ScanJob).filter(ScanJob.line_id == line.id).all()
        for job in jobs:
            db.query(Scan).filter(Scan.job_id == job.id).delete()
            db.query(Reconciliation).filter(Reconciliation.job_id == job.id).delete()
            db.query(ReconciliationQueue).filter(ReconciliationQueue.job_id == job.id).delete()

        # Delete jobs
        db.query(ScanJob).filter(ScanJob.line_id == line.id).delete()

        # Delete assignments
        db.query(Assignment).filter(Assignment.line_id == line.id).delete()

        # Delete the line
        db.delete(line)

        # Audit log
        actor_name = session.get(SESSION_TL_KEY, {}).get('display_name', 'Unknown')
        audit = AuditLog(
            actor=actor_name,
            action='LINE_DELETE',
            entity='LINE',
            entity_id=line.id,
            payload_json=json.dumps({
                'location': location,
                'warehouse': warehouse,
                'line_code': line_code,
                'deleted_by': 'manager' if (is_manager or is_manager_delete) else 'tl',
                'original_tl': assignment.tl_name
            })
        )
        db.add(audit)

        db.commit()
        return jsonify({'ok': True})

    except Exception as e:
        db.rollback()
        return jsonify({'error': str(e)}), 500
    finally:
        db.close()

@app.route('/api/mdf/fresh', methods=['POST'])
def api_create_fresh_mdf():
    """Create a fresh MDF.xlsx file (backup old one if exists)"""
    if not require_tl():
        return jsonify({"error": "TL authentication required"}), 401

    try:
        # Backup existing MDF if it exists
        backup_name = None
        if os.path.exists(MDF_PATH):
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            backup_name = f"MDF_backup_{timestamp}.xlsx"
            backup_path = os.path.join(EXPORTS_DIR, backup_name)

            # Copy the existing file as backup
            import shutil
            shutil.copy2(MDF_PATH, backup_path)

        # Create fresh MDF with just headers
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        ws.append(COLUMNS)
        wb.save(MDF_PATH)

        # Add audit log
        tl_session = session.get(SESSION_TL_KEY, {})
        audit = AuditLog(
            actor=tl_session.get("display_name", "TL"),
            action='FRESH_MDF_CREATED',
            entity='MDF',
            payload_json=json.dumps({"backup_created": backup_name})
        )

        db = SessionLocal()
        try:
            db.add(audit)
            db.commit()
        finally:
            db.close()

        return jsonify({
            "success": True,
            "message": "Fresh MDF created successfully",
            "backup": backup_name
        })

    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/api/logs/delete_all', methods=['DELETE'])
def api_delete_all_logs():
    """Delete all log entries (database and historical)"""
    if not require_tl():
        return jsonify({"error": "TL authentication required"}), 401

    data = request.get_json() or {}
    passcode = data.get('passcode', '')

    # Verify passcode
    if passcode != '240986':
        return jsonify({"error": "Invalid passcode"}), 403

    db = SessionLocal()
    try:
        # Get all completed jobs from database
        jobs = db.query(ScanJob).filter(
            ScanJob.status.in_(['submitted', 'variance_approved'])
        ).all()

        job_count = len(jobs)

        # Delete all database jobs
        for job in jobs:
            db.query(Scan).filter(Scan.job_id == job.id).delete()
            db.query(Reconciliation).filter(Reconciliation.job_id == job.id).delete()
            db.query(ReconciliationQueue).filter(ReconciliationQueue.job_id == job.id).delete()
            db.delete(job)

        # Create fresh Excel file (removes all historical data)
        with FileLock(LOCK_PATH, timeout=10):
            from openpyxl import Workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "Sheet1"
            ws.append(COLUMNS)  # Only headers, no data
            wb.save(MDF_PATH)
            wb.close()

        # Add audit log
        tl_session = session.get(SESSION_TL_KEY, {})
        audit = AuditLog(
            actor=tl_session.get("display_name", "TL"),
            action='ALL_LOGS_DELETE',
            entity='SCANJOB',
            payload_json=json.dumps({"deleted_count": job_count, "historical_deleted": True})
        )
        db.add(audit)

        db.commit()
        return jsonify({"success": True, "deleted_count": job_count})

    except Exception as e:
        db.rollback()
        return jsonify({"error": str(e)}), 500
    finally:
        db.close()

@app.route('/line-management')
def line_management():
    return render_template('line_management.html')

@app.route('/api/line-management/all')
def api_line_management_all():
    """Get all lines with their assignments for management - show all but allow editing only own lines"""
    if not require_tl():
        return jsonify({"ok": False, "reason": "unauthorized"}), 401

    db = SessionLocal()
    try:
        tl_session = session.get(SESSION_TL_KEY, {})
        tl_name = tl_session.get("display_name", "")
        is_manager = session.get('is_manager', False)
        tl_norm = _norm(tl_name)

        # Get all lines with assignments
        lines_with_assignments = db.query(Line, Assignment).join(
            Assignment,
            (Assignment.line_id == Line.id) & (Assignment.active == True)
        ).all()

        # Get pending requests for this TL or all if manager
        my_reqs = db.query(ReconciliationRequest).filter(
            ReconciliationRequest.status == 'pending'
        ).all()
        
        # Map line_id -> req for this TL identity
        req_by_line = {}
        for r in my_reqs:
            if is_manager or (r.tl_name_norm == tl_norm):
                req_by_line.setdefault(r.line_id, []).append(r)

        lines_data = []
        for line, assignment in lines_with_assignments:
            # Check if line has completed jobs
            completed_jobs = db.query(ScanJob).filter(
                ScanJob.line_id == line.id,
                ScanJob.status.in_(['submitted', 'variance_approved'])
            ).count()

            # Check if line has open jobs
            open_jobs = db.query(ScanJob).filter(
                ScanJob.line_id == line.id,
                ScanJob.status == 'open'
            ).count()

            # Determine status
            if completed_jobs > 0 and open_jobs == 0:
                status = 'completed'
                status_display = '✅ Completed'
            elif open_jobs > 0:
                status = 'active'
                status_display = '🔄 Active'
            else:
                status = 'not_started'
                status_display = '⏳ Not Started'

            # Check for pending requests
            reqs = req_by_line.get(line.id, [])
            pending_request = None
            if reqs:
                # Pick oldest pending
                pending = sorted(reqs, key=lambda x: x.created_at)[0]
                pending_request = {
                    "request_id": pending.id,
                    "requested_qty": int(pending.requested_qty or 0),
                    "requested_by": pending.requested_by,
                    "reason": pending.reason
                }

            line_data = {
                'id': line.id,
                'line_code': line.line_code,
                'location': line.location,
                'warehouse': line.warehouse,
                'target_qty': line.target_qty,
                'created_at': line.created_at.strftime('%Y-%m-%d %H:%M'),
                'updated_at': line.updated_at.strftime('%Y-%m-%d %H:%M'),
                'counter_name_1': assignment.counter_name_1 if assignment else None,
                'counter_name_2': assignment.counter_name_2 if assignment else None,
                'tl_name': assignment.tl_name if assignment else 'Not assigned',
                'can_edit': is_manager or (tl_name.lower() == assignment.tl_name.lower() if assignment else False),
                'status': status,
                'status_display': status_display,
                'completed_jobs': completed_jobs,
                'open_jobs': open_jobs,
                'pending_request': pending_request
            }
            lines_data.append(line_data)

        return jsonify({'ok': True, 'lines': lines_data, 'current_tl': tl_name, 'is_manager': is_manager})

    finally:
        db.close()

@app.route('/insights')
def insights():
    return render_template('insights.html')

@app.route('/api/counter/assignments')
def api_counter_assignments():
    """Get assignments for a specific counter"""
    counter_name = request.args.get('counter_name', '').strip()

    if not counter_name:
        return jsonify({"ok": False, "reason": "missing_counter"}), 400

    db = SessionLocal()
    try:
        print(f"DEBUG ASSIGNMENTS: Looking for assignments for counter: '{counter_name}'")
        
        # Find all assignments where this counter is assigned (case-insensitive)
        assignments = db.query(Assignment, Line).join(
            Line, Assignment.line_id == Line.id
        ).filter(
            Assignment.active == True
        ).all()

        print(f"DEBUG ASSIGNMENTS: Found {len(assignments)} total active assignments")

        counter_assignments = []
        counter_name_lower = counter_name.lower().strip()
        
        for assignment, line in assignments:
            print(f"DEBUG ASSIGNMENTS: Checking assignment - Line {line.line_code} at {line.location}/{line.warehouse}")
            print(f"DEBUG ASSIGNMENTS: Counter1: '{assignment.counter_name_1}', Counter2: '{assignment.counter_name_2}'")
            
            # Check if counter matches either position (case-insensitive)
            counter1_match = assignment.counter_name_1 and assignment.counter_name_1.lower().strip() == counter_name_lower
            counter2_match = assignment.counter_name_2 and assignment.counter_name_2.lower().strip() == counter_name_lower
            
            if not (counter1_match or counter2_match):
                print(f"DEBUG ASSIGNMENTS: Counter '{counter_name}' not assigned to line {line.line_code}")
                continue
                
            print(f"DEBUG ASSIGNMENTS: ✅ Counter '{counter_name}' IS assigned to line {line.line_code}")
            
            # Check if there's an active job for this line
            current_job = db.query(ScanJob).filter(
                ScanJob.line_id == line.id,
                ScanJob.status.in_(['open', 'locked_recon', 'variance_approved'])
            ).first()
            
            # Also check for completed jobs
            completed_jobs = db.query(ScanJob).filter(
                ScanJob.line_id == line.id,
                ScanJob.status == 'submitted'
            ).count()
            
            if current_job:
                print(f"DEBUG ASSIGNMENTS: Active job found for line {line.line_code} with status: {current_job.status}")
                counter_assignments.append({
                    'location': line.location,
                    'warehouse': line.warehouse,
                    'line_code': line.line_code,
                    'target_qty': line.target_qty,
                    'tl_name': assignment.tl_name,
                    'line_id': line.id,
                    'job_id': current_job.id,
                    'job_status': current_job.status
                })
            elif completed_jobs == 0:
                print(f"DEBUG ASSIGNMENTS: No jobs found for line {line.line_code} - creating assignment opportunity")
                # Include lines without any jobs - they need to be started
                counter_assignments.append({
                    'location': line.location,
                    'warehouse': line.warehouse,
                    'line_code': line.line_code,
                    'target_qty': line.target_qty,
                    'tl_name': assignment.tl_name,
                    'line_id': line.id,
                    'job_id': None,
                    'job_status': 'ready_to_start'
                })
            else:
                print(f"DEBUG ASSIGNMENTS: Line {line.line_code} has completed jobs but no active job - line finished")

        print(f"DEBUG ASSIGNMENTS: ✅ FINAL RESULT: Returning {len(counter_assignments)} assignments to frontend")
        for assignment in counter_assignments:
            print(f"DEBUG ASSIGNMENTS: ➤ {assignment['location']}/{assignment['warehouse']}/Line {assignment['line_code']} (Status: {assignment['job_status']})")
        
        return jsonify({
            "ok": True,
            "assignments": counter_assignments
        })

    finally:
        db.close()

@app.route('/api/insights/dashboard')
def api_insights_dashboard():
    """Get dashboard insights data for managers"""
    db = SessionLocal()
    try:
        # Get total lines
        total_lines = db.query(Line).count()

        # Get job statistics
        active_jobs = db.query(ScanJob).filter(ScanJob.status == 'open').count()
        completed_jobs = db.query(ScanJob).filter(ScanJob.status.in_(['submitted', 'variance_approved'])).count()

        # Get total scans
        total_scans = db.query(func.sum(Scan.qty)).scalar() or 0

        # Get location data
        location_data = []
        for location in LOCATIONS:
            lines_count = db.query(Line).filter(Line.location == location).count()
            active_jobs_count = db.query(ScanJob).join(Line).filter(
                Line.location == location,
                ScanJob.status == 'open'
            ).count()

            location_data.append({
                'location': location,
                'lines': lines_count,
                'activeJobs': active_jobs_count
            })

        # Get status data
        status_data = []
        statuses = ['open', 'submitted', 'variance_approved']
        for status in statuses:
            count = db.query(ScanJob).filter(ScanJob.status == status).count()
            if count > 0:
                status_data.append({
                    'status': status.replace('_', ' ').title(),
                    'count': count
                })

        # Get TL performance
        tl_performance = []
        assignments = db.query(Assignment).filter(Assignment.active == True).all()
        tl_stats = {}

        for assignment in assignments:
            tl_name = assignment.tl_name
            if tl_name not in tl_stats:
                tl_stats[tl_name] = {
                    'name': tl_name,
                    'linesManaged': 0,
                    'activeJobs': 0,
                    'completedJobs': 0,
                    'totalScans': 0
                }

            tl_stats[tl_name]['linesManaged'] += 1

            # Get jobs for this line
            line_jobs = db.query(ScanJob).filter(ScanJob.line_id == assignment.line_id).all()
            for job in line_jobs:
                if job.status == 'open':
                    tl_stats[tl_name]['activeJobs'] += 1
                elif job.status in ['submitted', 'variance_approved']:
                    tl_stats[tl_name]['completedJobs'] += 1

                # Get scans for this job
                job_scans = db.query(func.sum(Scan.qty)).filter(Scan.job_id == job.id).scalar() or 0
                tl_stats[tl_name]['totalScans'] += job_scans

        tl_performance = list(tl_stats.values())

        # Get active lines detail
        active_lines = []
        lines_with_jobs = db.query(Line, ScanJob).join(ScanJob).filter(
            ScanJob.status == 'open'
        ).all()

        for line, job in lines_with_jobs:
            # Get assignment
            assignment = db.query(Assignment).filter(
                Assignment.line_id == line.id,
                Assignment.active == True
            ).first()

            # Get scanned count
            scanned = db.query(func.sum(Scan.qty)).filter(Scan.job_id == job.id).scalar() or 0

            assigned_counters = []
            if assignment:
                if assignment.counter_name_1:
                    assigned_counters.append(assignment.counter_name_1)
                if assignment.counter_name_2:
                    assigned_counters.append(assignment.counter_name_2)

            active_lines.append({
                'lineCode': line.line_code,
                'location': line.location,
                'warehouse': line.warehouse,
                'target': line.target_qty,
                'scanned': int(scanned),
                'status': job.status,
                'assignedCounters': ', '.join(assigned_counters) if assigned_counters else 'Not assigned'
            })

        return jsonify({
            'ok': True,
            'totalLines': total_lines,
            'activeJobs': active_jobs,
            'completedJobs': completed_jobs,
            'totalScans': int(total_scans),
            'locationData': location_data,
            'statusData': status_data,
            'tlPerformance': tl_performance,
            'activeLines': active_lines
        })

    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)}), 500
    finally:
        db.close()

@app.route('/health')
def health():
    return jsonify({'ok': True})

if __name__ == '__main__':
    with app.app_context():
        init_db()
        ensure_mdf()
    print("\n🚀 DSV STOCK COUNT - Line-Based Stock Count App Started!")
    print("Access the app at: http://0.0.0.0:5000")

    app.run(host='0.0.0.0', port=5000, debug=True)
